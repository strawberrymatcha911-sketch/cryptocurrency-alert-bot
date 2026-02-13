import requests
from openpyxl import Workbook, load_workbook
import time
from datetime import datetime

# ================= Telegram & Crypto API Setup =================
BOT_TOKEN = "8320044697:AAGueY738b3l4l3mtZKrgin34-Hip4Ac6Bw"
CHAT_ID = "5029207640"
API_KEY = "0b14c1a43cab9703b36444e5aa748a6f0fa3dc9a7e6b07008d2c9a5abd9cd0e5"
cryptos = ["BTC", "ETH", "SOL", "MANA", "SAND", "DOGE"]
currency = "USD"
crypto_url = "https://min-api.cryptocompare.com/data/pricemulti"
headers = {"authorization": f"Apikey {API_KEY}"}
file_name = "LiveCryptoPrices.xlsx"

def send_alert(crypto, alert_type, target_price, live_price):
    """
    Send Telegram alert with clickable CryptoCompare link
    """
    crypto_link = f"https://www.cryptocompare.com/coins/{crypto.lower()}/overview/USD"

    message = (
        f"*{alert_type} ALERT*\n"
        f"Crypto: {crypto}\n"
        f"Target Price: {target_price} USD\n"
        f"Current Price: {live_price} USD\n\n"
        f"[Buy / Sell on CryptoCompare]({crypto_link})"
    )

    url = f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage"
    payload = {
        "chat_id": CHAT_ID,
        "text": message,
        "parse_mode": "Markdown"
    }

    try:
        requests.post(url, data=payload, timeout=10)
        print("ALERT SENT TO TELEGRAM")
    except Exception as e:
        print("FAILED TO SEND ALERT:", e)

def init_excel():
    """Initialize Excel workbook and worksheet"""
    try:
        wb = load_workbook(file_name)
        ws = wb.active
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.title = "Live Prices"
        ws.append(["Time", "Crypto", "Price (USD)"])
        wb.save(file_name)
    return wb, ws

while True:
    print('--------------------------------------------------------------')
    print('    Cryptocurrency Portfolio Application Main Menu')
    print('--------------------------------------------------------------')
    print('7. CryptoAlertBot')
    print('E. Exit Main Menu')
    print('--------------------------------------------------------------')
    main_menu = input("Select an option: ")

        # function 7
    if main_menu == "7":
         # User Input 
        while True:
            crypto_name = input("Enter cryptocurrency name (BTC/ETH/SOL/MANA/SAND/DOGE): ").upper()
            if crypto_name in cryptos:
                break
            print("Invalid crypto. Choose from:", ", ".join(cryptos))

        while True:
            try:
                buy_price = float(input(f"Enter BUY target price for {crypto_name}: "))
                sell_price = float(input(f"Enter SELL target price for {crypto_name}: "))
                break
            except ValueError:
                print("Please enter numeric values for prices.")

        # Excel Setup 
        wb, ws = init_excel()

        print("\nMonitoring started... Press Ctrl+C to stop.\n")

        # Monitoring Loop 
        try:
            while True:
                response = requests.get(
                    crypto_url,
                    params={"fsyms": crypto_name, "tsyms": currency},
                    headers=headers
                )

                data = response.json()
                live_price = data.get(crypto_name, {}).get(currency)

                if live_price is None:
                    print("API error, retrying in 30s...")
                    time.sleep(30)
                    continue

                now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                ws.append([now, crypto_name, live_price])
                wb.save(file_name)

                print(f"{now} | {crypto_name} = {live_price}")

                if abs(live_price - buy_price) <= 200:
                    send_alert(crypto_name, "BUY", buy_price, live_price)

                if abs(live_price - sell_price) <= 200:
                    send_alert(crypto_name, "SELL", sell_price, live_price)

                time.sleep(30)

        except KeyboardInterrupt:
            print("\nMonitoring stopped. Returning to main menu.")

    # E: Exit 
    elif main_menu.upper() == "E":
        print("Exiting program...")
        break

    else:
        print("Invalid selection. Try again.")
